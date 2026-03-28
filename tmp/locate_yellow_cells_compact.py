import openpyxl
from openpyxl.styles import PatternFill

file_path = r'C:\Users\Enrique Saavedra\Documents\Programa ACT\public\New Contract Modification Log amarillo.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['CML Report']

yellows = []
for r in range(1, 200):
    for c in range(1, 30):
        cell = ws.cell(row=r, column=c)
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == 'FFFFFF00':
            l = str(ws.cell(row=r, column=max(1, c-1)).value)
            r_val = str(ws.cell(row=r, column=min(ws.max_column, c+1)).value)
            t = str(ws.cell(row=max(1, r-1), column=c).value)
            yellows.append(f"{cell.coordinate}: L='{l}', R='{r_val}', T='{t}', V='{cell.value}'")

print("\n".join(yellows))
