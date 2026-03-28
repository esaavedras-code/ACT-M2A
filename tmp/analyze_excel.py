import openpyxl
import json
import sys

wb = openpyxl.load_workbook(r'C:\Users\Enrique Saavedra\Documents\Programa ACT\public\AC017630 EWO 08 (M) New Contract Modification Log.xlsx')
print(wb.sheetnames)
ws = wb.active
out = {}
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            out[cell.coordinate] = str(cell.value)

with open(r'C:\Users\Enrique Saavedra\Documents\Programa ACT\tmp\excel_py.json', 'w') as f:
    json.dump(out, f, indent=2)
