import openpyxl
from openpyxl.styles import PatternFill

file_path = r'C:\Users\Enrique Saavedra\Documents\Programa ACT\public\New Contract Modification Log amarillo.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['CML Report']

print(f"Yellow cells in CML Report (Rows 1-65):")
for r in range(1, 66):
    for c in range(1, 30):
        cell = ws.cell(row=r, column=c)
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb == 'FFFFFF00':
            left = ws.cell(row=r, column=max(1, c-1)).value
            right = ws.cell(row=r, column=min(ws.max_column, c+1)).value
            top = ws.cell(row=max(1, r-1), column=c).value
            print(f"{cell.coordinate} (R{r}C{c}): Val='{cell.value}', Left='{left}', Right='{right}', Top='{top}'")
