import openpyxl
import json

try:
    wb = openpyxl.load_workbook('Documentos/Comparacion_PS.xlsx', data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"Sheet: {sheet_name}")
        ws = wb[sheet_name]
        for row in list(ws.iter_rows(values_only=True))[:30]:
            if any(row):
                print(row[:10])
except Exception as e:
    print("Error:", e)
