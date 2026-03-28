import os
import json
from dotenv import load_dotenv
from groq import Groq
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Cargar variables de entorno
load_dotenv(".env.local")
api_key = os.getenv("GROQ_API_KEY")

def format_excel_report(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACT-96 Informe Generado"
    
    # Estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="004B87") # Azul institucional
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    
    # Encabezado principal
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "INFORME DIARIO DE INSPECCIÓN (ACT-96)"
    cell.font = Font(bold=True, size=14)
    cell.alignment = center_align
    
    # Escribir Datos Básicos
    fields = [
        ("NÚM. DE PROYECTO (1)", data.get("numero_proyecto", "")),
        ("NOMBRE DE PROYECTO (2)", data.get("nombre_proyecto", "")),
        ("MUNICIPIO (3)", data.get("municipio", "")),
        ("CONTRATISTA (4)", data.get("contratista", "")),
        ("FECHA (5)", data.get("fecha", "")),
        ("CLIMA (9)", data.get("clima", "")),
        ("ADMINISTRADOR (24)", data.get("administrador", "")),
        ("PUESTO (25)", data.get("puesto", "")),
    ]
    
    row = 3
    for label, val in fields:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=str(val)).alignment = left_align
        row += 1
        
    row += 1
    # Actividades
    ws.merge_cells(f"A{row}:D{row}")
    cell = ws.cell(row=row, column=1, value="TRABAJO EJECUTADO / ACTIVIDADES (15)")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    row += 1
    
    for act in data.get("actividades", []):
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=1, value="Actividad:").font = bold_font
        ws.cell(row=row, column=2, value=str(act)).alignment = left_align
        row += 1
        
    row += 1
    # Equipo
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws.cell(row=row, column=1, value="EQUIPO")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    
    # Personal
    ws.merge_cells(f"C{row}:D{row}")
    cell = ws.cell(row=row, column=3, value="PERSONAL")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    row += 1
    
    equipos = data.get("equipos", [])
    personal = data.get("personal", [])
    max_lines = max(len(equipos), len(personal))
    for i in range(max_lines):
        ws.merge_cells(f"A{row}:B{row}")
        if i < len(equipos):
            ws.cell(row=row, column=1, value=str(equipos[i])).alignment = left_align
            
        ws.merge_cells(f"C{row}:D{row}")
        if i < len(personal):
            ws.cell(row=row, column=3, value=str(personal[i])).alignment = left_align
        row += 1
        
    # Guardar
    wb.save(output_path)
    print(f"Reporte de Excel ACT-96 generado exitosamente en: {output_path}")

def main():
    if not api_key:
        print("ERROR: No se encontró GROQ_API_KEY en .env.local")
        return
        
    docs_dir = r"C:\Users\Enrique Saavedra\Documents\Programa ACT\Documentos"
    inspeccion_file = os.path.join(docs_dir, "ACT-96 Inspeccion.txt")
    instrucciones_file = os.path.join(docs_dir, "ACT-96 Instrucciones.txt")
    
    # Leer PDFs (texto)
    with open(inspeccion_file, "r", encoding="utf-8") as f:
        inspeccion_text = f.read()
    with open(instrucciones_file, "r", encoding="utf-8") as f:
        instrucciones_text = f.read()
        
    prompt = f"""Eres un experto inspector residente que automatiza reportes.
Tengo un informe de inspección diario lleno en un PDF pero escaneado (a continuación como Texto del Informe) y un manual de instrucciones de cómo se debe llenar el ACT-96 (a continuación como Texto Instructivo).
Por favor, analiza el informe y extrae de forma PERFECTAMENTE estructurada en JSON estrictamente con la siguiente estructura, sacando TODA la información pertinente basada en las instrucciones.

Formato requerido JSON:
{{
  "numero_proyecto": "...", 
  "nombre_proyecto": "...",
  "municipio": "...",
  "contratista": "...",
  "fecha": "...",
  "clima": "...",
  "actividades": ["actividad 1", "actividad 2..."],
  "equipos": ["equipo 1", "equipo 2..."],
  "personal": ["empleado 1", "empleado 2..."],
  "administrador": "...",
  "puesto": "..."
}}

--- TEXTO INSTRUCTIVO ---
{instrucciones_text}

--- TEXTO DEL INFORME (A EXTRAER) ---
{inspeccion_text}

NOTA: Si en las actividades hay detalles como (2 limpiezas, ok), extrae la línea completa para no perder información.
IMPORTANTE: Devuelve ÚNICAMENTE el código JSON. No agregues comillas invertidas extra ni palabras antes o después.
"""

    print("Analizando datos con Groq AI...")
    client = Groq(api_key=api_key)
    
    response = client.chat.completions.create(
        messages=[{"role": "user", "content": prompt}],
        model="llama-3.3-70b-versatile",
        temperature=0.1,
    )
    
    output_text = response.choices[0].message.content.strip()
    # Limpiar si empieza con ```json
    if output_text.startswith("```json"):
        output_text = output_text[7:]
    if output_text.startswith("```"):
        output_text = output_text[3:]
    if output_text.endswith("```"):
        output_text = output_text[:-3]
        
    try:
        data = json.loads(output_text.strip())
        print("Datos extraídos correctamente.")
    except Exception as e:
        print("Error al decodificar JSON. El LLM devolvió:")
        print(output_text)
        return
        
    json_path = os.path.join(docs_dir, "ACT-96_Inspeccion_Datos.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
        
    excel_path = os.path.join(docs_dir, "ACT-96_Reporte_Generado_AI.xlsx")
    format_excel_report(data, excel_path)

if __name__ == "__main__":
    main()
